import cv2
import os
from openpyxl import load_workbook
from pipeline.preprocessing import detect_voter_boxes
from pipeline.ocr_engine import image_array_to_text
from pipeline.parser import parse_voter_fields

wb = load_workbook('output/voter_output.xlsx')
ws = wb.active

# Find rows with leading colon in names
bad_rows = []
for row in range(9, ws.max_row + 1):
    name = ws.cell(row, 3).value
    if name and str(name).startswith(':'):
        bad_rows.append(row)

print(f"Found {len(bad_rows)} names with leading colon:")
print()

# Map row to card index: Each page has 30 cards, first data row is 9
temp_images = sorted([f for f in os.listdir('images') if f.endswith('.jpg')])

for bad_row in bad_rows[:2]:  # Check first 2 bad records
    card_global_idx = bad_row - 9
    page_idx = card_global_idx // 30
    card_idx = card_global_idx % 30
    
    excel_name = ws.cell(bad_row, 3).value
    
    if page_idx < len(temp_images):
        img_path = os.path.join('images', temp_images[page_idx])
        img = cv2.imread(img_path)
        if img is not None:
            boxes = detect_voter_boxes(img, min_width=400, min_height=150)
            if card_idx < len(boxes):
                x, y, w, h = boxes[card_idx]
                crop = img[y:y+h, x:x+w]
                text = image_array_to_text(crop, psm=6)
                
                print(f"Row {bad_row}: {repr(excel_name)}")
                print(f"  Page {temp_images[page_idx]} Card {card_idx+1}")
                print(f"  OCR text: {repr(text[:300])}")
                print()
