import cv2
import os
from openpyxl import load_workbook
from pipeline.preprocessing import detect_voter_boxes
from pipeline.ocr_engine import image_array_to_text

wb = load_workbook('output/voter_output.xlsx')
ws = wb.active
temp_images = sorted([f for f in os.listdir('images') if f.endswith('.jpg')])

# Check row 318 (which has ': Divyamani M')
excel_row = 318
excel_name = ws.cell(excel_row, 3).value

# Calculate which page and card this should be
# First data row is 9
# Row 318 means it's the (318-9) = 309th record (0-indexed: 308)
card_global_idx = excel_row - 9 - 1
page_idx = card_global_idx // 30
card_idx = card_global_idx % 30

print(f"Excel row {excel_row}:")
print(f"  Name from Excel: {repr(excel_name)}")
print(f"  Global card index: {card_global_idx}")
print(f"  Should be on page {page_idx} (image: {temp_images[page_idx]})")
print(f"  Card position on page: {card_idx}")
print()

# Now check the actual image
if page_idx < len(temp_images):
    img_path = os.path.join('images', temp_images[page_idx])
    img = cv2.imread(img_path)
    if img is not None:
        boxes = detect_voter_boxes(img, min_width=400, min_height=150)
        print(f"  Total cards on page: {len(boxes)}")
        if card_idx < len(boxes):
            x, y, w, h = boxes[card_idx]
            crop = img[y:y+h, x:x+w]
            text = image_array_to_text(crop, psm=6)
            print(f"  OCR text: {repr(text[:200])}")
            
            # Also check the identity by looking at unique names
            # Let's check card positions directly
            print(f"\n  Checking first and last cards on this page:")
            for check_idx in [0, len(boxes)-1]:
                if check_idx < len(boxes):
                    x2, y2, w2, h2 = boxes[check_idx]
                    crop2 = img[y2:y2+h2, x2:x2+w2]
                    text2 = image_array_to_text(crop2, psm=6)
                    # Extract name from OCR
                    import re
                    m = re.search(r'(?<!\w)Name\s*[=:]\s*(\S[^\n=:]*?)', text2, re.IGNORECASE)
                    ocr_name = m.group(1).strip() if m else "N/A"
                    excel_row_for_card = excel_row - card_idx + check_idx
                    excel_name_for_card = ws.cell(excel_row_for_card, 3).value
                    print(f"    Card {check_idx}: Excel row {excel_row_for_card} has {repr(excel_name_for_card)}, OCR has {repr(ocr_name)}")
