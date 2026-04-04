# pipeline/exporter.py
import pandas as pd
from typing import List, Dict, Optional
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
 
VOTER_HEADERS = [
    'Serial Number',
    'EPIC Number',
    'Name',
    'Relative Name',
    'Relation Type',
    'House Number',
    'Age',
    'Gender',
    'Record Type',
]


def deduplicate_records(records: List[Dict]) -> tuple:
    """
    Remove duplicate records that were extracted twice (once as Voter, once as Amendment).
    
    Duplicates are identified by matching (EPIC Number, Name, House Number).
    When duplicates exist, keep the "Voter" record type and discard others.
    
    Returns: (deduplicated_list, removed_count)
    """
    if not records:
        return records, 0
    
    # Build a key from record identity fields
    def get_record_key(rec):
        epic = str(rec.get('EPIC Number', '')).strip()
        name = str(rec.get('Name', '')).strip()
        house = str(rec.get('House Number', '')).strip()
        # Use non-empty fields only
        parts = [p for p in [epic, name, house] if p and p.lower() != 'nan']
        return tuple(parts) if parts else None
    
    # Group records by key
    seen = {}  # key -> (index, record)
    deduplicated = []
    removed_count = 0
    
    for idx, rec in enumerate(records):
        key = get_record_key(rec)
        if key is None:
            # No identifying info, keep it
            deduplicated.append(rec)
            continue
        
        if key not in seen:
            # First occurrence of this record - keep it
            seen[key] = (idx, rec)
            deduplicated.append(rec)
        else:
            # Duplicate found
            prev_idx, prev_rec = seen[key]
            prev_type = prev_rec.get('Record Type', 'Voter')
            curr_type = rec.get('Record Type', 'Voter')
            
            # Prefer to keep "Voter" type; if both or neither are voters, keep first
            if prev_type == 'Voter' or curr_type != 'Voter':
                # Keep the previous one, discard this one
                removed_count += 1
            else:
                # Replace with current one (which is Voter type)
                deduplicated[-1] = rec
                seen[key] = (idx, rec)
                removed_count += 1  # Still removing one
    
    if removed_count > 0:
        print(f"Deduplicated: removed {removed_count} duplicate records")
    
    return deduplicated, removed_count
 
 
def _to_voter_df(records: List[Dict]) -> pd.DataFrame:
    """Normalize arbitrary record dicts into the fixed voter export schema."""
    df = pd.DataFrame(records)
    return df.reindex(columns=VOTER_HEADERS)
 
 
def save_to_csv(records: List[Dict], out_path: str):
    """Write records as plain CSV with stable voter column ordering."""
    df = _to_voter_df(records)
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    df.to_csv(out_path, index=False, encoding='utf-8')
    return out_path
 
 
def save_to_excel(records: List[Dict], out_path: str):
    """Write records as a basic Excel file without custom styling."""
    df = _to_voter_df(records)
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
    df.to_excel(out_path, index=False)
    return out_path
 
 
def save_to_formatted_excel(records: List[Dict], out_path: str, metadata: Optional[Dict] = None):
    """
    Export voter records to a clean, properly formatted Excel file.
    
    If metadata is provided, writes polling station details at the top before the voter data table.

    Layout with metadata:
      Rows 1-5 — metadata section (Polling Station details)
      Row 6 — blank separator
      Row 7 — merged title: "Electoral Roll Data — N voters"
      Row 8 — bold column headers
      Row 9+ — data rows
    
    Layout without metadata:
      Row 1 — merged title: "Electoral Roll Data — N voters"
      Row 2 — bold column headers
      Row 3+ — data rows
    """
    if metadata is None:
        metadata = {
            'polling_station_number': '',
            'polling_station_name': '',
            'polling_station_address': '',
            'block': '',
            'ward': '',
        }
    
    headers = VOTER_HEADERS
    df = _to_voter_df(records)
 
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)
 
    wb = Workbook()
    ws = wb.active
    ws.title = 'Voter Data'
 
    num_cols = len(headers)
    last_col = get_column_letter(num_cols)
    
    current_row = 1
    metadata_fill = PatternFill("solid", fgColor="D9E1F2")
    metadata_font = Font(bold=True, size=11)
    
    # Write metadata section if any metadata is present
    has_metadata = any([
        metadata.get('polling_station_number'),
        metadata.get('polling_station_name'),
        metadata.get('polling_station_address'),
        metadata.get('block'),
        metadata.get('ward'),
    ])
    
    if has_metadata:
        # Metadata header label
        meta_header = ws.cell(current_row, 1, "POLLING STATION DETAILS")
        meta_header.font = Font(bold=True, size=12, color="FFFFFF")
        meta_header.fill = PatternFill("solid", fgColor="203864")
        meta_header.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
        ws.row_dimensions[current_row].height = 18
        current_row += 1
        
        # Metadata rows
        metadata_fields = [
            ('Polling Station Number', 'polling_station_number'),
            ('Polling Station Name', 'polling_station_name'),
            ('Polling Station Address', 'polling_station_address'),
            ('Block', 'block'),
            ('Ward', 'ward'),
        ]
        
        for label, key in metadata_fields:
            value = metadata.get(key, '')
            label_cell = ws.cell(current_row, 1, f"{label}:")
            label_cell.font = metadata_font
            label_cell.fill = metadata_fill
            label_cell.alignment = Alignment(horizontal='left')
            
            value_cell = ws.cell(current_row, 2, value if value else "")
            value_cell.fill = metadata_fill
            value_cell.alignment = Alignment(horizontal='left', wrap_text=True)
            ws.merge_cells(f"B{current_row}:{last_col}{current_row}")
            ws.row_dimensions[current_row].height = 18
            current_row += 1
        
        # Blank separator row
        current_row += 1
 
    # Row for title (Electoral Roll Data title)
    title_cell = ws.cell(current_row, 1, f"Electoral Roll Data — {len(records)} voters")
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.fill = title_fill
    ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
    ws.row_dimensions[current_row].height = 22
    current_row += 1
 
    # Column headers row
    header_fill = PatternFill("solid", fgColor="2E75B6")
    for c, header in enumerate(headers, 1):
        cell = ws.cell(current_row, c, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill
    current_row += 1
 
    # Data rows
    if not df.empty:
        for row in df.itertuples(index=False):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(current_row, c_idx, value)
                cell.alignment = Alignment(horizontal='left')
            # Light alternating row shading for readability
            if current_row % 2 == 0:
                row_fill = PatternFill("solid", fgColor="EBF3FB")
                for c_idx in range(1, num_cols + 1):
                    ws.cell(current_row, c_idx).fill = row_fill
            current_row += 1
 
    # Column widths
    column_widths = {
        'A': 12,  # Serial Number
        'B': 13,  # EPIC Number
        'C': 25,  # Name
        'D': 25,  # Relative Name
        'E': 14,  # Relation Type
        'F': 15,  # House Number
        'G': 8,   # Age
        'H': 10,  # Gender
        'I': 14,  # Record Type
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
 
    wb.save(out_path)
    return out_path